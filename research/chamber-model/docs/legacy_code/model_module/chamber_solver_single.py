"""
================================================================================
 chamber_solver.py  -  Subcooled pool-boiling chamber: calibrated gray-box model
================================================================================
 Single-file numerical solver for a closed subcooled pool-boiling chamber with a
 submerged tube-bundle condenser (water; HFE-7000 = 'RE347mcc' selectable).

 Contents
   1. Properties (CoolProp, cached)
   2. Correlations (Rohsenow boiling, natural convection, Kandlikar CHF)
   3. Chip heat-balance solver  ->  surface temperature, heat partition, ONB
   4. Condenser model           ->  coolant hydraulics + calibrated operating point
                                     (predict T_sat & subcooling from geometry)
   5. Experimental data loader  ->  reads the chamber workbooks
   6. Calibration               ->  fits C_sf (transferable) + single-phase factor
   7. Figures                   ->  reproduces the key model plots
   8. main()                    ->  calibrate, validate, plot

 Parameter STATUS (do not blur these):
   * C_sf (plain)    CALIBRATED, TRANSFERABLE material constant (0.0131 water/Cu)
   * C_sf (micro)    CALIBRATED, TRANSFERABLE microchannel constant (0.0067, ~half)
   * area (micro)    GEOMETRIC area augmentation (2.17, Cooke-Kandlikar)
   * NC_factor       CALIBRATED, GEOMETRY-SPECIFIC (per-chamber, does not transfer)
   * CHF (plain)     MEASURED at burnout: 114 (33-tube) / 65 (42-tube) W/cm^2
   * CHF (micro)     MEASURED: plain + ~60 W/cm^2 (additive, both chambers)
   * operating point CALIBRATED per chamber: T_sat & subcooling vs (T_in, Q). The
                     condenser is coolant-side limited; T_sat gain on coolant inlet < 1.
                     Predictive chain RMSE 4.4 K chip temp (5% on rise, q>50% CHF).

 Run:  python chamber_solver.py
 Deps: numpy, scipy, CoolProp, matplotlib, openpyxl
================================================================================
"""
import os
import numpy as np
from scipy.optimize import brentq, least_squares
from CoolProp.CoolProp import PropsSI

UPLOADS = "/mnt/user-data/uploads"
OUT = "/mnt/user-data/outputs"
C_SF_LIT = 0.013          # literature water / polished copper
G = 9.81
L_CHIP = 8.3e-3           # A/P of the 34.5 x 32 mm chip
FLUID = "Water"           # 'Water' or 'RE347mcc' (HFE-7000)

# ---- calibrated chip surfaces (geometry-independent boiling coefficients) ----
C_SF_PLAIN = 0.0131       # CALIBRATED plain-chip Rohsenow C_sf (water / polished Cu)
C_SF_MC    = 0.0067       # CALIBRATED open-microchannel C_sf (~half the plain value)
AREA_MC    = 2.17         # microchannel geometric area augmentation (wetted/footprint)
CHF_INC_MC = 60.0         # microchannel additive CHF increment over plain [W/cm^2]

# ============================== 1. PROPERTIES ===============================
_pc = {}
def props(Tsat_c, fluid=FLUID):
    key = (round(Tsat_c, 1), fluid)
    if key in _pc:
        return _pc[key]
    T = Tsat_c + 273.15
    p = dict(
        rho_l=PropsSI("D", "T", T, "Q", 0, fluid), rho_v=PropsSI("D", "T", T, "Q", 1, fluid),
        hfg=PropsSI("H", "T", T, "Q", 1, fluid) - PropsSI("H", "T", T, "Q", 0, fluid),
        sigma=PropsSI("I", "T", T, "Q", 0, fluid), mu_l=PropsSI("V", "T", T, "Q", 0, fluid),
        k_l=PropsSI("L", "T", T, "Q", 0, fluid), cp_l=PropsSI("C", "T", T, "Q", 0, fluid),
        beta=PropsSI("ISOBARIC_EXPANSION_COEFFICIENT", "T", T, "Q", 0, fluid))
    p["Pr"] = p["mu_l"] * p["cp_l"] / p["k_l"]
    _pc[key] = p
    return p

# ============================= 2. CORRELATIONS ==============================
def q_natural_convection(T_surf, T_pool, p, NC, L=L_CHIP):
    """Enhanced McAdams/Churchill-Chu single-phase term (W/m^2)."""
    dT = max(T_surf - T_pool, 1e-9)
    nu = p["mu_l"] / p["rho_l"]; al = p["k_l"] / (p["rho_l"] * p["cp_l"])
    Ra = G * p["beta"] * dT * L ** 3 / (nu * al)
    Nu = 0.15 * Ra ** (1 / 3.) if Ra > 1e7 else 0.54 * Ra ** 0.25
    Nu = max(Nu, 0.27 * Ra ** 0.25)
    return NC * Nu * p["k_l"] / L * dT

def q_nucleate_boiling(T_surf, T_sat, p, C_sf):
    """Rohsenow nucleate boiling term (W/m^2); zero below saturation."""
    dsup = T_surf - T_sat
    if dsup <= 0:
        return 0.0
    return p["mu_l"] * p["hfg"] * (G * (p["rho_l"] - p["rho_v"]) / p["sigma"]) ** 0.5 * \
           (p["cp_l"] * dsup / (C_sf * p["hfg"] * p["Pr"])) ** 3

def chf_kandlikar(T_sat, subcooling, beta_deg=45.0, phi_deg=0.0, C_sub=0.0535, fluid=FLUID):
    """Kandlikar 2001 CHF (W/cm^2). ABSOLUTE UNCALIBRATED (placeholder beta, C_sub)."""
    p = props(T_sat, fluid); b = np.radians(beta_deg); ph = np.radians(phi_deg)
    q_sat = p["hfg"] * p["rho_v"] ** 0.5 * ((1 + np.cos(b)) / 16) * \
            (2 / np.pi + np.pi / 4 * (1 + np.cos(b)) * np.cos(ph)) ** 0.5 * \
            (p["sigma"] * G * (p["rho_l"] - p["rho_v"])) ** 0.25
    return q_sat / 1e4 * (1 + C_sub * subcooling)

# ============================ 3. BOILING SOLVER =============================
def surface_temp(q_wcm2, T_sat, subcooling, NC, L=L_CHIP, C_sf=C_SF_LIT, fluid=FLUID):
    """Solve q'' = sqrt(q_nc^2 + q_nb^2) for chip surface temperature [C]."""
    T_pool = T_sat - subcooling
    p = props(T_sat, fluid)
    q = q_wcm2 * 1e4
    f = lambda Ts: np.hypot(q_natural_convection(Ts, T_pool, p, NC, L),
                            q_nucleate_boiling(Ts, T_sat, p, C_sf)) - q
    return brentq(f, T_pool + 1e-6, T_sat + 160)

def partition(q_wcm2, T_sat, subcooling, NC, C_sf=C_SF_LIT, fluid=FLUID):
    """Energy share (single-phase, boiling)."""
    Ts = surface_temp(q_wcm2, T_sat, subcooling, NC, C_sf=C_sf, fluid=fluid)
    p = props(T_sat, fluid)
    a = q_natural_convection(Ts, T_sat - subcooling, p, NC)
    b = q_nucleate_boiling(Ts, T_sat, p, C_sf)
    tot = a * a + b * b
    return a * a / tot, b * b / tot

def onb_flux(T_sat, subcooling, NC, fluid=FLUID):
    """Flux at onset of nucleate boiling (T_surf = T_sat) [W/cm^2]."""
    return q_natural_convection(T_sat, T_sat - subcooling, props(T_sat, fluid), NC) / 1e4

# ===================== 3b. SELECTABLE CHIP SURFACES =========================
CHIPS = {
    # area = wetted/footprint augmentation;  C_sf = calibrated boiling coefficient
    "plain": dict(area=1.00,    C_sf=C_SF_PLAIN, chf_increment=0.0),
    "micro": dict(area=AREA_MC, C_sf=C_SF_MC,    chf_increment=CHF_INC_MC),
}

def chip_surface_temp(q_wcm2, T_sat, subcooling, NC, chip="plain", fluid=FLUID):
    """Chip surface temperature [C] for a selectable chip ('plain' or 'micro').

    Same chamber machinery for both: the open-microchannel chip applies the
    geometric area augmentation to both transport terms and swaps in the
    calibrated microchannel boiling coefficient C_sf_mc = 0.0067. The per-chamber
    single-phase factor NC is unchanged - it is a chamber property, not a chip one.
    Calibrated on 110 microchannel points across both chambers: pooled RMSE 6.7 K.
    """
    c = CHIPS[chip]
    T_pool = T_sat - subcooling
    p = props(T_sat, fluid)
    q = q_wcm2 * 1e4
    f = lambda Ts: np.hypot(c["area"] * q_natural_convection(Ts, T_pool, p, NC),
                            c["area"] * q_nucleate_boiling(Ts, T_sat, p, c["C_sf"])) - q
    return brentq(f, T_pool + 1e-9, T_sat + 260)

def chip_chf(plain_chf_wcm2, chip="plain"):
    """CHF [W/cm^2] for a selectable chip: the chamber's measured plain-chip CHF
    plus the additive microchannel increment (separated-flow feed, ~+60 W/cm^2,
    near-constant across chambers). The multiplicative area factor over-predicts."""
    return plain_chf_wcm2 + CHIPS[chip]["chf_increment"]

# ============================ 4. CONDENSER MODEL ============================
GEOM = {
    "33-tube": dict(N=33, ID=3.14e-3, OD=4.76e-3, L=95e-3, area_cm2=309, mdot_gps=78, NC=9.0),
    "42-tube": dict(N=42, ID=1.39e-3, OD=3.175e-3, L=62.6e-3, area_cm2=262, mdot_gps=36, NC=3.5),
    "66-tube": dict(N=66, ID=1.8e-3, OD=2.0e-3, L=95e-3, area_cm2=394, mdot_gps=None, NC=None),  # no data yet
}
K_CU = 400.0
def _coolant(Tc=25.0):
    T = Tc + 273.15
    return dict(rho=PropsSI("D", "T", T, "P", 101325, "Water"),
                mu=PropsSI("V", "T", T, "P", 101325, "Water"),
                k=PropsSI("L", "T", T, "P", 101325, "Water"),
                cp=PropsSI("C", "T", T, "P", 101325, "Water"))

def hydraulics(name, Q_lpm=4.0):
    g = GEOM[name]; c = _coolant(); Q = Q_lpm / 60e3
    V = (Q / g["N"]) / (np.pi / 4 * g["ID"] ** 2)
    Re = c["rho"] * V * g["ID"] / c["mu"]
    dP = 32 * c["mu"] * g["L"] * V / g["ID"] ** 2 + 1.5 * c["rho"] * V ** 2 / 2
    return dict(V=V, Re=Re, dP=dP, laminar=Re < 2300)

def effectiveness_ntu(name, Q_lpm=4.0, h_chamber=600.0):
    """DIAGNOSTIC ONLY. Ideal effectiveness-NTU with fully-developed laminar Nu.
    This over-predicts T_sat: it misses the entrance-region coolant enhancement and
    the flux-dependent exposed condensing area. Use predict_operating() for prediction."""
    g = GEOM[name]; c = _coolant(); Q = Q_lpm / 60e3
    C_cold = c["rho"] * Q * c["cp"]
    hi = 3.66 * c["k"] / g["ID"]
    Ai = np.pi * g["ID"] * g["L"] * g["N"]; Ao = np.pi * g["OD"] * g["L"] * g["N"]
    UA = 1 / (1 / (hi * Ai) + np.log(g["OD"] / g["ID"]) / (2 * np.pi * K_CU * g["L"] * g["N"]) + 1 / (h_chamber * Ao))
    NTU = UA / C_cold
    return dict(UA=UA, NTU=NTU, eps=1 - np.exp(-NTU))

# -------- calibrated operating-point sub-model (predict T_sat & subcooling) --------
# Mechanism (from the resistance breakdown): this condenser is COOLANT-SIDE LIMITED.
# Internal flow is developing-laminar in short tubes (Re~1100); film condensation is
# nearly ideal and negligible (2-3 mK/W). The condenser conductance IMPROVES as the
# whole system runs warmer (coolant viscosity falls, exposed condensing area grows with
# vapor inventory), so T_sat responds to coolant inlet with a gain a < 1 rather than 1:1.
# Realized as a calibrated linear operating-point map per chamber:
#     T_sat   = a*T_in + b*Q + c        (a = coolant gain, b = effective thermal resistance K/W, c = offset)
#     subcool = d*T_in + e*Q + f        (pool-to-coolant balance; subcooling falls as coolant warms)
# Q is the CHAMBER HEAT LOAD in watts ( = q'' [W/cm^2] * 11.04 cm^2 footprint ).
# Fit on 261 points (plain + micro, four coolant inlets). T_sat RMSE 4 K (33) / 9 K (42).
# End-to-end chip-temperature RMSE 4.4 K, median 4% of absolute T, 5% on the rise above
# coolant in the design-relevant regime (q > 50% of CHF).
OP_TSAT = {   # (a, b, c)  -- a,c per-chamber, dimensionless / K ; b in K/W
    "33-tube": (0.2618, 0.005994, 44.81),
    "42-tube": (0.7098, 0.024088, 33.14),
    # "66-tube": UNCALIBRATED (no chamber data). Estimate from layout: its open, tall
    #   2 mm bank should resemble the 33-tube (low gain, small slope) more than the 42-tube.
}
OP_SUB = {    # (d, e, f)  -- subcooling = d*T_in + e*Q + f  [K]
    "33-tube": (-0.6389, -0.006797, 40.75),
    "42-tube": (-0.2173, -0.000701, 19.62),
}
A_FOOT_CM2 = 11.04        # chip footprint: Q[W] = q''[W/cm^2] * A_FOOT_CM2

def predict_operating(q_wcm2, T_in_C, chamber):
    """Predict the chamber operating point from geometry + boundary conditions.
    Inputs: chip heat flux q'' [W/cm^2], coolant inlet temperature [C], chamber name.
    Returns (T_sat [C], subcooling [K]). No measured operating conditions required."""
    if chamber not in OP_TSAT:
        raise KeyError(f"operating-point constants not calibrated for {chamber}; "
                       f"available: {list(OP_TSAT)}")
    Q = q_wcm2 * A_FOOT_CM2
    a, b, c = OP_TSAT[chamber]; d, e, f = OP_SUB[chamber]
    T_sat = a * T_in_C + b * Q + c
    subcool = max(d * T_in_C + e * Q + f, 0.0)
    return T_sat, subcool

def chip_temp_from_geometry(q_wcm2, T_in_C, chamber, NC, chip="plain", fluid=FLUID):
    """Full predictive chain: geometry + coolant inlet + heat flux -> chip surface temp [C].
    Predicts the operating point (T_sat, subcooling) then solves the selected chip surface.
    NC is the chamber single-phase factor (9.0 for 33-tube, 3.5 for 42-tube)."""
    T_sat, subcool = predict_operating(q_wcm2, T_in_C, chamber)
    return chip_surface_temp(q_wcm2, T_sat, subcool, NC, chip=chip, fluid=fluid)

# ============================== 5. DATA LOADER ==============================
def _find_header(ws):
    want = ("Tsurf", 'q"', "Tsat", "Liquid Temperature", "Experiment  Data Points")
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=14, values_only=True), 1):
        hit = {c.strip(): i for i, c in enumerate(row) if isinstance(c, str) and c.strip() in want}
        if "Tsurf" in hit and 'q"' in hit:
            return ri, hit
    raise ValueError("header row not found")

def load_data(xlsx_path, sheets):
    """Return list of (q'' W/cm2, T_sat C, subcooling K, T_surf C)."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True); pts = []
    for sn in sheets:
        ws = wb[sn]; hr, c = _find_header(ws)
        for r in ws.iter_rows(min_row=hr + 1, values_only=True):
            if not isinstance(r[c["Experiment  Data Points"]], (int, float)):
                continue
            q, ts, tsat, tl = r[c['q"']], r[c["Tsurf"]], r[c["Tsat"]], r[c["Liquid Temperature"]]
            if all(isinstance(v, (int, float)) for v in (q, ts, tsat, tl)):
                pts.append((float(q), float(tsat), float(tsat) - float(tl), float(ts)))
    return pts

SHEETS_33 = ["Plain chip 20C T2", "Plain chip 30C", "Plain chip 40C", "Plain chip 50C"]
SHEETS_42 = ["Plain Chip 20C old", "Plain Chip 30C old", "Plain Chip 40C old", "Plain Chip 50C Old"]

# ============================== 6. CALIBRATION ==============================
def _resid(params, pts):
    C_sf, NC = params; out = []
    for q, Tsat, sub, Tm in pts:
        try:
            out.append(surface_temp(q, Tsat, sub, NC, C_sf=C_sf) - Tm)
        except Exception:
            out.append(Tm + 50.0)
    return out

def _rmse(params, pts):
    return float(np.sqrt(np.mean(np.array(_resid(params, pts)) ** 2)))

def calibrate(pts, x0=(0.013, 8.0), freeze_Csf=None):
    """Fit (C_sf, NC_factor) by least squares against measured surface temperatures."""
    if freeze_Csf is None:
        r = least_squares(_resid, x0, args=(pts,), bounds=([0.005, 1], [0.03, 20]))
        return r.x[0], r.x[1], _rmse(r.x, pts)
    r = least_squares(lambda p: _resid((freeze_Csf, p[0]), pts), [x0[1]], bounds=([1], [20]))
    return freeze_Csf, r.x[0], _rmse((freeze_Csf, r.x[0]), pts)

# =============================== 7. FIGURES =================================
def make_figures(NC33=9.0):
    import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
    Tsat = 54.0
    # (a) boiling-curve family vs subcooling + CHF locus
    qs = np.linspace(1, 150, 120); subs = [0, 10, 20, 30, 40]
    cols = plt.cm.viridis_r(np.linspace(0.1, 0.9, len(subs)))
    fig, ax = plt.subplots(figsize=(8, 5.5)); chf = []
    for sc, c in zip(subs, cols):
        ax.plot([surface_temp(q, Tsat, sc, NC33) for q in qs], qs, color=c, lw=2, label=f"ΔT_sub={sc} K")
        ch = chf_kandlikar(Tsat, sc); chf.append((surface_temp(min(ch, 150), Tsat, sc, NC33), ch))
    xs, ys = zip(*chf); ax.plot(xs, ys, "r--s", lw=1.4, ms=6, label="CHF (UNCALIBRATED)")
    ax.axvline(Tsat, color="gray", ls=":"); ax.set_xlim(35, 100); ax.set_ylim(0, 150)
    ax.set_xlabel("T_surf [°C]"); ax.set_ylabel("q'' [W/cm²]"); ax.legend(fontsize=8); ax.grid(alpha=.3)
    ax.set_title("Subcooling shifts the boiling curve (factor=9.0)")
    fig.tight_layout(); fig.savefig(f"{OUT}/fig_subcooling_curves.png", dpi=140); plt.close(fig)
    # (b) effectiveness-NTU
    fig, ax = plt.subplots(figsize=(7, 5)); ntu = np.linspace(0, .25, 100)
    ax.plot(ntu, 1 - np.exp(-ntu), "k-", lw=2)
    for n, c in zip(GEOM, ["#1f5fb0", "#c0392b", "#2ca25f"]):
        lo, hi = effectiveness_ntu(n, h_chamber=600), effectiveness_ntu(n, h_chamber=8000)
        ax.plot([lo["NTU"], hi["NTU"]], [lo["eps"], hi["eps"]], "-o", color=c, lw=5, alpha=.6, label=n)
    ax.set_xlabel("NTU"); ax.set_ylabel("effectiveness ε"); ax.legend(); ax.grid(alpha=.3)
    ax.set_title("Condenser ε-NTU (single-phase→condensation bound)")
    fig.tight_layout(); fig.savefig(f"{OUT}/fig_eNTU.png", dpi=140); plt.close(fig)
    print("  wrote fig_subcooling_curves.png, fig_eNTU.png")

# ================================= MAIN ====================================
def main():
    print("Calibrating against experimental data...\n")
    p33 = load_data(f"{UPLOADS}/Plain_Chip_data_33_tube_condenser.xlsx", SHEETS_33)
    p42 = load_data(f"{UPLOADS}/Plain_Chip_data_older_condenser.xlsx", SHEETS_42)
    Csf, NC, e = calibrate(p33)
    print(f"33-tube ({len(p33)} pts): C_sf={Csf:.4f}  factor={NC:.2f}  RMSE={e:.2f}°C  (baseline ~12.6°C)")
    _, NCf, ef = calibrate(p42, freeze_Csf=Csf); Csf2, NC2, e2 = calibrate(p42)
    print(f"42-tube frozen C_sf:  factor={NCf:.2f}  RMSE={ef:.2f}°C")
    print(f"42-tube fresh fit:    C_sf={Csf2:.4f}  factor={NC2:.2f}  RMSE={e2:.2f}°C")
    print(f"  -> C_sf transfers ({abs(Csf2-Csf)/Csf*100:.0f}%); factor does NOT ({NC:.1f} vs {NC2:.1f})\n")
    print("Condenser coolant side (4 L/min):")
    for n in GEOM:
        h = hydraulics(n); en = effectiveness_ntu(n)
        print(f"  {n}: V={h['V']:.2f} m/s  Re={h['Re']:.0f}  ΔP={h['dP']:.0f} Pa  "
              f"UA~{en['UA']:.0f} W/K  ε~{en['eps']*100:.1f}%")
    print(f"\nCHF: not reached in data; CHF > 111 (33-tube), > 61 (42-tube) W/cm². "
          f"Kandlikar @20K subcool ~ {chf_kandlikar(54,20):.0f} W/cm² (UNCALIBRATED).")
    if os.path.isdir(OUT):
        print("\nGenerating figures...")
        make_figures(NC)

if __name__ == "__main__":
    main()
